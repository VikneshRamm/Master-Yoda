import base64
import openpyxl
import json

def parse_feedback_excel(file_path):
    try:
        file_path = base64.b64decode(file_path.encode('utf-8')).decode('utf-8')
        # Load the workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        all_sheets_data = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_content = {}
            current_key = None
            
            # We start from row 2 to skip the main title "ENGINEER-LEAD FEEDBACK DOCUMENT"
            # max_col=2 because we only care about columns A and B
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                key_cell, value_cell = row
                
                # If Column A has text, it's a new category (e.g., "Appreciations")
                if key_cell:
                    # Create a clean key name (e.g., "Action Items" -> "action_items")
                    current_key = key_cell.strip().lower().replace(" ", "_")
                    sheet_content[current_key] = []
                    
                    if value_cell:
                        sheet_content[current_key].append(str(value_cell).strip())
                
                # If Column A is empty but Column B has text, it's a continuation point
                elif current_key and value_cell:
                    sheet_content[current_key].append(str(value_cell).strip())
            
            # Store this sheet's data using the sheet name as the key
            all_sheets_data[sheet_name] = sheet_content
            
        json_output = json.dumps(all_sheets_data)
        return json_output
    except Exception as e:
        print(f"Error parsing Excel file: {str(e)}")
        return ""
